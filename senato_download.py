# -*- coding: utf-8 -*-
"""
Created on Sun Nov 12 21:33:15 2023

@author: edobo
"""

from bs4 import BeautifulSoup
import requests
import pandas as pd
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
import os
import time
import re
import webbrowser
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry


# Load the Excel workbook
wb = openpyxl.load_workbook(r"C:\Users\edobo\OneDrive\Desktop\Thesis\Medieval Guilds\Data\EDITED_db\Merge\check_senato.xlsx")
sheet = wb.active
destination_folder = r"C:\Users\edobo\OneDrive\Desktop\Thesis\Medieval Guilds\Data\senato"
URL = []

name = pd.read_excel(r"C:\Users\edobo\OneDrive\Desktop\Thesis\Medieval Guilds\Data\EDITED_db\Merge\check_senato.xlsx")
names = name['Place']
# Iterate through rows in the sheet
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    # Iterate through cells in the row
    for cell in row:
        # Check if the cell has a hyperlink
        if cell.hyperlink:
            # Print the hyperlink address
            URL.append(cell.hyperlink.target)
            
def make_request_with_retry(url, retry_strategy=None):
    session = requests.Session()

    if retry_strategy:
        adapter = HTTPAdapter(max_retries=retry_strategy)
        session.mount("http://", adapter)
        session.mount("https://", adapter)

    return session.get(url, timeout=10)

# Example retry configuration
retry_strategy = Retry(
    total=3,
    status_forcelist=[429, 500, 502, 503, 504],
    method_whitelist=["HEAD", "GET", "OPTIONS"]
)


n = 0

for url in URL:
    name = names[n]
    n = n + 1
    url_def = url + '&Click=C1256DF7003BC91C.85be9e93542e661dc1256ddc003b74a3/$Body/2.3416'
    pdf_name = str(name) + '.pdf'
    pdf_path = os.path.join(destination_folder, pdf_name)

    try:
        pdf_response = make_request_with_retry(url_def, retry_strategy)
        pdf_response.raise_for_status()  # Raise an error for bad responses

        soup = BeautifulSoup(pdf_response.content, 'html.parser')
        script = soup.find('script')

        if script:
            script_text = script.text
            pattern = r"window.location='(.*?)'"
            link = re.findall(pattern, script_text)
            link = str(link)
            link = link.replace("[", "").replace("]", "").replace("'", "")

            pdf_response = make_request_with_retry(link, retry_strategy)
            
            with open(pdf_path, 'wb') as pdf_file:
                pdf_file.write(pdf_response.content)

            print(f"PDF {name} downloaded successfully.")
        else:
            print(f"No script found for PDF {name}.")

    except requests.exceptions.RequestException as e:
        print(f"Error fetching PDF {name}: {e}")

gilde = []    

for url in URL:
    try:
        # Use the make_request_with_retry function
        response = make_request_with_retry(url, retry_strategy=retry_strategy)
        response.raise_for_status()  # Raise HTTPError for bad responses

        # Parse HTML content with BeautifulSoup
        soup = BeautifulSoup(response.content, 'html.parser')

        # Find all anchor tags (links) in the HTML
        denominazione_ente_td = soup.find_all('table', attrs={'width': '480'})

        for row in denominazione_ente_td:
            tr = row.find_all('tr')
            denominazione_ente = ""
            localita = ""
            for thing in tr:
                cells = thing.find_all('td', {'bgcolor': '#FFFF7F'})
                for cell in cells:
                    # Check if the cell contains "DENOMINAZIONE ENTE"
                    if "DENOMINAZIONE ENTE" in cell.get_text():
                        # Extract the value from the next cell
                        value_cell = cell.find_next('td', {'bgcolor': '#FFFFC2'})
                        denominazione_ente = value_cell.get_text(strip=True)

                    if 'LOCALITA' in cell.get_text():
                        value_cell = cell.find_next('td', {'bgcolor': '#FFFFC2'})
                        localita = value_cell.get_text(strip=True)

            gilde.append((localita, denominazione_ente))

    except requests.exceptions.RequestException as e:
        print(f"Error fetching data from {url}: {e}")


gilde = pd.DataFrame(gilde, columns = ['Place', 'Name of the guild'])      
#gilde = gilde.drop_duplicates()
gilde.to_excel(r"C:\Users\edobo\OneDrive\Desktop\Thesis\Medieval Guilds\Data\EDITED_db\Merge\check_senato_output.xlsx")

                

